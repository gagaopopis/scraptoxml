# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
import datetime

class ItemExcelPipeline(object):
    column_names = {'_id':'ID (Артикул)*', '_group_id':'Групповой идентификатор',
               'typePrefix':'Тип', 'model':"Наименование*", "description":"Описание",
               "url":"Веб-адрес", "picture":'Адрес картинки', "price":"Цена (руб)*", "old_price":"Старая цена (руб)",
               "p_series":"Линейка","p_filler":"Наполнение",
               "p_material":"Материал каркаса","p_color":"Цвет",
               "p_wood_type":"Цвет дерева", "p_transformation":"Механизм трансформации",
               "p_suspension":"Пружинный блок", "p_modular":"Модульный",
               "p_box":"Ящик для белья", "p_armrests":"Подлокотники",
               "p_corner":"Угловой","p_width":"Ширина","p_depth":"Глубина",
                     "p_height":"Высота", "p_sleep_width":"Ширина спального места",
                     "p_sleep_length": "Длина спального места"
                     }
    columns_sofa = ['_id','_available', 'model', "price", "old_price", "description",
     "picture", "picture_1", "picture_2", "picture_3", "picture_4",
     "p_color", "p_transform", "p_transformation",
     "p_upholstery", "p_cloth", "country_of_origin",
     "p_depth", "p_width","p_height", "p_sleep_width","p_sleep_length",
     "p_box", "p_armrests", "p_bar", "p_usb", "p_remcover", "p_pillows"]
    columns_corner = ['_id','_available', 'model', "price", "old_price", "description",
     "picture", "picture_1", "picture_2", "picture_3", "picture_4",
     "p_corner","p_color", "p_transform", "p_transformation",
     "p_upholstery", "p_cloth", "country_of_origin",
     "p_depth", "p_width","p_height", "p_sleep_width","p_sleep_length",
     "p_box", "p_armrests", "p_bar", "p_usb", "p_remcover", "p_pillows"]
    columns_chair = ['_id','_available', 'model', "price", "old_price", "description",
     "picture", "picture_1", "picture_2", "picture_3", "picture_4",
     "p_chair","p_color", "p_transform", "p_transformation",
     "p_upholstery", "p_cloth", "country_of_origin",
     "p_depth", "p_width","p_height", "p_sleep_width","p_sleep_length",
     "p_box", "p_armrests", "p_bar", "p_usb", "p_remcover", "p_pillows"]
    
    columns_puf = ['_id','_available', 'model', "price", "old_price", "description",
     "picture", "picture_1", "picture_2", "picture_3", "picture_4",
     "p_color", "p_upholstery", "p_cloth", "country_of_origin",
     "p_depth", "p_width","p_height",
     "p_box", "p_remcover"]
    columns_couch = ['_id','_available', 'model', "price", "old_price", "description",
     "picture", "picture_1", "picture_2", "picture_3", "picture_4",
     "p_color", "p_transform", "p_transformation",
     "p_upholstery", "p_cloth", "country_of_origin",
     "p_depth", "p_width","p_height", "p_sleep_width","p_sleep_length",
     "p_box", "p_armrests", "p_remcover", "p_pillows"]

    columns = {"Прямые диваны":columns_sofa,
      "Угловые диваны":columns_corner,
       "Модульные диваны": columns_sofa,
       "Кресла и мешки": columns_chair,
       "Пуфы и банкетки":columns_puf,
       "Кушетки":columns_couch }

        
    def __init__(self, xlsxfle):
        self.xlsxfile = xlsxfle

    @classmethod
    def from_crawler(cls, crawler):
        return cls(xlsxfle=crawler.settings.get('XLSXFILE'))

    def open_spider(self, spider):
        self.wb = load_workbook(filename="template.xlsx")

    def close_spider(self, spider):
        self.wb.save(self.xlsxfile)

    def process_item(self, item, spider):
        ws_name = item.get('typePrefix')
        columns = self.columns[ws_name]
        ws = self.wb[ws_name]
        row = ws.max_row + 1
        for c in columns:
            col = columns.index(c) + 1
            v = item.get(c)
            ws.cell(row, col, value=v)
        return item
    """
    row = self.ws.max_row + 1
    for c in self.columns:
        col = self.columns.index(c)+1
        v = item.get(c)
        if isinstance(v, list):
            v = v[0]
        self.ws.cell(row, col,value=v)
    return item
    """


class ItemXMLPipeline(object):

    def __init__(self, xmlfile):
        self.xmlfile = xmlfile
        self.tree = None

    @classmethod
    def from_crawler(cls, crawler):
        return cls(xmlfile=crawler.settings.get('XMLFILE'))

    def open_spider(self, spider):
        try:
            self.tree = ET.parse(self.xmlfile)
        except FileNotFoundError:
            self.tree = ET.ElementTree()
            root = ET.Element('yml_catalog', attrib={'date': "{0:%Y-%m-%d %H:%M}".format(datetime.datetime.now())})
            shop = ET.SubElement(root, 'shop')
            self.tree._setroot(root)

    def close_spider(self, spider):
        self.tree.write(self.xmlfile, encoding='utf-8', xml_declaration=True)

    def process_item(self, item, spider):
        shop = self.tree.find('shop')
        offer = ET.SubElement(shop, 'offer', attrib={"id":item['_id'],
                                                     'type': "vendor.model",
                                                     'group_id': item.get('_group_id', default=''),
                                                     'available':item.get('_available',default='false')
                                                     }
                              )
        for key in (x for x in item.keys() if not x.startswith("_")):
            value = item.get(key)
            if isinstance(value, list):
                for l in value:
                    p = ET.SubElement(offer, key, attrib=item.fields[key])
                    p.text = l
            elif key.startswith("p_"):
                p = ET.SubElement(offer, 'param', attrib=item.fields[key])
                p.text = value
            else:
                p = ET.SubElement(offer, key, attrib=item.fields[key])
                p.text = value

        return item
